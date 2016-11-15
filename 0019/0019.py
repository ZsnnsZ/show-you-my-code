#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from xml.dom.minidom import Document
from openpyxl import load_workbook, Workbook

readname = 'numbers.xlsx'
savename = 'numbers.xml'


def load_data(readname):
    wb = load_workbook(filename=readname)
    sheet = wb.get_active_sheet()
    data = []
    for i in range(3):
        data.append([])
        for j in range(3):
            data[i].append(sheet.cell(row=i + 1, column=j + 1).value)
    # print(data)
    return data


def write_xlsx_to_xml(data, savename):
    # 创建dom文档
    doc = Document()
    # 创建根节点
    root = doc.createElement('root')
    # 创建student节点
    student = doc.createElement('numbers')
    # 根节点插入dom树
    doc.appendChild(root)
    # 将student子节点加入根节点
    root.appendChild(student)
    # 加入注释
    comment = doc.createComment('\n    数字信息\n')
    student.appendChild(comment)
    # 加入内容
    text_node = doc.createTextNode(str(data))
    student.appendChild(text_node)
    # 保存为xml文件
    with open(savename, 'w', encoding='utf-8') as f:
        doc.writexml(f, addindent='  ', newl='\n', encoding='utf-8')

if __name__ == '__main__':
    data = load_data(readname)
    write_xlsx_to_xml(data, savename)
