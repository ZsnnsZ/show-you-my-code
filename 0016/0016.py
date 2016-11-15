#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import json

readname = 'numbers.txt'
savename = 'numbers.xlsx'


def save_xlsx(readname):
    with open(readname, 'r', encoding='utf-8') as f:
        data = json.load(f, encoding='utf-8')
    print(data)
    wb = Workbook()
    ws = wb.active
    ws.title = 'numbers'
    for i in range(1, len(data) + 1):
        for j in range(len(data[i - 1])):
            ws.cell(row=i, column=j + 1).value = data[i - 1][j]
    wb.save(filename=savename)

if __name__ == '__main__':
    save_xlsx(readname)
