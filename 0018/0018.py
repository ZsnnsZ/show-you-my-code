#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from xml.dom.minidom import Document
from openpyxl import load_workbook, Workbook

readname = 'city.xlsx'
savename = 'city.xml'


def load_data(readname):
    wb = load_workbook(filename=readname)
    sheet = wb.get_active_sheet()
    data = {}
    for i in range(1, 4):
        x = sheet.cell(row=i, column=2).value
        data[str(sheet.cell(row=i, column=1).value)] = str(x)
    return data


def write_xlsx_to_xml(datadict, savename):
    # 创建dom文档
    doc = Document()
    # 创建根节点
    root = doc.createElement('root')
    # 创建student节点
    student = doc.createElement('city')
    # 根节点插入dom树
    doc.appendChild(root)
    # 将student子节点加入根节点
    root.appendChild(student)
    # 加入注释
    comment = doc.createComment('\n    城市信息\n')
    student.appendChild(comment)
    # 加入内容
    text_node = doc.createTextNode(str(datadict))
    student.appendChild(text_node)
    # 保存为xml文件
    with open(savename, 'w', encoding='utf-8') as f:
        doc.writexml(f, addindent='  ', newl='\n', encoding='utf-8')

if __name__ == '__main__':
    datadict = load_data(readname)
    write_xlsx_to_xml(datadict, savename)
