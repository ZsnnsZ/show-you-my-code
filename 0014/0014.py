#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import json

readname = 'students.txt'
savename = 'students.xlsx'


def save_xlsx(readname):
    with open(readname, 'r', encoding='utf-8') as f:
        json_data = json.load(f, encoding='utf-8')
    print(json_data)
    wb = Workbook()
    ws = wb.active
    ws.title = 'student'
    for i in range(1, len(json_data) + 1):
        ws.cell(row=i, column=1).value = i
        for j in range(len(json_data[str(i)])):
            ws.cell(row=i, column=j + 2).value = json_data[str(i)][j]
    wb.save(filename=savename)

if __name__ == '__main__':
    save_xlsx(readname)
